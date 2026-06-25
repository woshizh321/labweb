"""Parse an uploaded Excel/CSV into the 44 Kawasaki_IVIG model variables.
Single-patient autofill: reads a 2-column (label, value) or wide (header row +
first data row) sheet, maps each recognized label/key to a model feature, and
returns a de-identified {feature: value} dict. Unrecognized columns (e.g.
patient_id / visit_id / name) are silently ignored, so identifiers never pass through.
"""
from __future__ import annotations

import csv
import io
from typing import Any

KINDS: dict[str, str] = {
    "age_years": "number",
    "sex": "sex",
    "max_temp_pre_ivig": "number",
    "fever_days_ivig": "number",
    "rash_days_ivig": "number",
    "rash": "binary",
    "conjunctival_injection": "binary",
    "strawberry_tongue": "binary",
    "cracked_lips": "binary",
    "oral_mucosal_change": "binary",
    "cervical_lymphadenopathy": "binary",
    "extremity_edema": "binary",
    "periungual_desquamation": "binary",
    "extremity_change": "binary",
    "wbc": "number",
    "neutrophil_percent": "number",
    "lymphocyte_percent": "number",
    "monocyte_percent": "number",
    "hemoglobin": "number",
    "platelet": "number",
    "crp": "number",
    "esr": "number",
    "pct": "number",
    "ferritin": "number",
    "alt": "number",
    "ast": "number",
    "albumin": "number",
    "total_bilirubin": "number",
    "direct_bilirubin": "number",
    "creatinine": "number",
    "urea_nitrogen": "number",
    "uric_acid": "number",
    "sodium": "number",
    "potassium": "number",
    "pt": "number",
    "aptt": "number",
    "fibrinogen": "number",
    "cd4_t_count": "number",
    "cd8_t_count": "number",
    "cd4_cd8_ratio": "number",
    "cd19_b_count": "number",
    "ldh": "number",
    "ck_mb": "number",
    "ntprobnp": "number",
}

# feature -> accepted header labels (model key, form label, dictionary label, English)
ALIASES: dict[str, list[str]] = {
    "age_years": ["Age", "age_years", "年龄"],
    "sex": ["Sex", "sex", "性别"],
    "max_temp_pre_ivig": ["IVIG前最高体温", "Max temperature before IVIG", "max_temp_pre_ivig"],
    "fever_days_ivig": ["Fever duration at IVIG", "IVIG时发热天数", "fever_days_ivig"],
    "rash_days_ivig": ["IVIG时皮疹天数", "Rash duration at IVIG", "rash_days_ivig"],
    "rash": ["Rash", "rash", "皮疹"],
    "conjunctival_injection": ["Conjunctival injection", "conjunctival_injection", "结膜充血"],
    "strawberry_tongue": ["Strawberry tongue", "strawberry_tongue", "草莓舌"],
    "cracked_lips": ["Cracked lips", "cracked_lips", "口唇皲裂"],
    "oral_mucosal_change": ["Oral mucosal change", "oral_mucosal_change", "口腔粘膜改变", "口腔黏膜改变"],
    "cervical_lymphadenopathy": ["Cervical lymphadenopathy", "cervical_lymphadenopathy", "颈部淋巴结肿大"],
    "extremity_edema": ["Extremity edema", "extremity_edema", "手足硬肿"],
    "periungual_desquamation": ["Periungual desquamation", "periungual_desquamation", "肢端脱皮"],
    "extremity_change": ["Extremity change", "extremity_change", "指趾端改变"],
    "wbc": ["IVIG前白细胞", "White blood cell count", "wbc"],
    "neutrophil_percent": ["IVIG前中性粒细胞百分比", "Neutrophil percentage", "neutrophil_percent"],
    "lymphocyte_percent": ["IVIG前淋巴细胞百分比", "Lymphocyte percentage", "lymphocyte_percent"],
    "monocyte_percent": ["Monocyte percentage", "monocyte_percent", "单核细胞百分数"],
    "hemoglobin": ["Hemoglobin", "IVIG前血红蛋白", "hemoglobin"],
    "platelet": ["IVIG前血小板", "Platelet count", "platelet"],
    "crp": ["C-reactive protein", "IVIG前C反应蛋白", "crp"],
    "esr": ["Erythrocyte sedimentation rate", "esr", "血沉"],
    "pct": ["Procalcitonin", "pct", "降钙素原"],
    "ferritin": ["Ferritin", "ferritin", "铁蛋白"],
    "alt": ["Alanine aminotransferase", "alt", "谷丙转氨酶"],
    "ast": ["Aspartate aminotransferase", "ast", "谷草转氨酶"],
    "albumin": ["Albumin", "albumin", "白蛋白"],
    "total_bilirubin": ["Total bilirubin", "total_bilirubin", "总胆红素"],
    "direct_bilirubin": ["Direct bilirubin", "direct_bilirubin", "直接胆红素"],
    "creatinine": ["Creatinine", "creatinine", "肌酐"],
    "urea_nitrogen": ["Urea nitrogen", "urea_nitrogen", "尿素氮"],
    "uric_acid": ["Uric acid", "uric_acid", "尿酸"],
    "sodium": ["Sodium", "sodium", "钠"],
    "potassium": ["Potassium", "potassium", "钾"],
    "pt": ["PT", "Prothrombin time", "pt", "凝血酶原时间"],
    "aptt": ["APTT", "Activated partial thromboplastin time", "aptt", "活化部分凝血活酶时间"],
    "fibrinogen": ["Fibrinogen", "fibrinogen", "纤维蛋白原"],
    "cd4_t_count": ["CD4+ T-cell count", "CD4+T细胞绝对值", "cd4_t_count"],
    "cd8_t_count": ["CD8+ T-cell count", "CD8+T细胞绝对值", "cd8_t_count"],
    "cd4_cd8_ratio": ["CD4/CD8 ratio", "CD4/CD8比值", "cd4_cd8_ratio"],
    "cd19_b_count": ["CD19+ B-cell count", "CD19+B细胞绝对值", "cd19_b_count"],
    "ldh": ["Lactate dehydrogenase", "ldh", "乳酸脱氢酶"],
    "ck_mb": ["CK-MB", "Creatine kinase-MB", "ck_mb", "肌酸激酶同工酶"],
    "ntprobnp": ["NT-proBNP", "ntprobnp"],
}


class ImportParseError(ValueError):
    """Raised when the uploaded file cannot be parsed at all."""


def _norm(s: Any) -> str:
    t = str(s).strip().lower().replace(" ", "").replace("　", "")
    # unify common variant character and drop any parenthetical unit/suffix
    t = t.replace("粘", "黏")  # 粘 -> 黏
    for op, cl in (("(", ")"), ("（", "）")):
        while op in t and cl in t and t.index(op) < t.index(cl):
            t = t[: t.index(op)] + t[t.index(cl) + 1 :]
    return t


_LABEL_TO_FEATURE = {_norm(a): f for f, al in ALIASES.items() for a in al}

_TRUE = {"1", "1.0", "yes", "y", "true", "是", "有", "阳性"}
_FALSE = {"0", "0.0", "no", "n", "false", "否", "无", "阴性"}
_MALE = {"male", "m", "男", "男性", "1"}
_FEMALE = {"female", "f", "女", "女性", "0", "2"}


def _is_blank(v: Any) -> bool:
    return v is None or str(v).strip() in {"", "nan", "na", "none", "未填写", "-"}


def _coerce(feature: str, raw: Any):
    kind = KINDS[feature]
    s = str(raw).strip()
    if kind == "sex":
        t = s.lower()
        if t in _MALE:
            return "male"
        if t in _FEMALE:
            return "female"
        return None
    if kind == "binary":
        t = s.lower()
        if t in _TRUE:
            return 1
        if t in _FALSE:
            return 0
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _rows_from_bytes(filename: str, content: bytes) -> list[list[Any]]:
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        try:
            from openpyxl import load_workbook
        except Exception as exc:  # pragma: no cover
            raise ImportParseError("xlsx support unavailable") from exc
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        return [list(r) for r in ws.iter_rows(values_only=True)]
    # CSV / TSV fallback
    text = content.decode("utf-8-sig", errors="replace")
    sep = "\t" if (name.endswith(".tsv") or "\t" in text.splitlines()[0:1][0:1] and "," not in text[:200]) else ","
    return [row for row in csv.reader(io.StringIO(text), delimiter=sep)]


def parse_upload(filename: str, content: bytes) -> dict[str, Any]:
    rows = [r for r in _rows_from_bytes(filename, content) if r and any(not _is_blank(c) for c in r)]
    if not rows:
        raise ImportParseError("empty file")

    pairs: list[tuple[Any, Any]] = []
    width = max(len(r) for r in rows)
    if width <= 3:
        # vertical: (label, value) per row
        for r in rows:
            if len(r) >= 2:
                pairs.append((r[0], r[1]))
    else:
        # wide: header row + first data row with a real value
        header = rows[0]
        data = next((r for r in rows[1:] if any(not _is_blank(c) for c in r)), [])
        for i, h in enumerate(header):
            pairs.append((h, data[i] if i < len(data) else None))

    out: dict[str, Any] = {}
    for label, value in pairs:
        if _is_blank(label) or _is_blank(value):
            continue
        feature = _LABEL_TO_FEATURE.get(_norm(label))
        if not feature or feature in out:
            continue
        coerced = _coerce(feature, value)
        if coerced is not None:
            out[feature] = coerced
    return out
